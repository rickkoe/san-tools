import xml.etree.cElementTree as ET 
import os
import csv
import itertools
import sys
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, colors, Font, Color
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

if len(sys.argv) == 2:
    # If an argument exists, the first argument is customer name
    customer_name = sys.argv[1]
else:
    customer_name = input('Enter customer name: ')
import importlib
config = importlib.import_module(f'data.{customer_name}.config')
now = datetime.now()
date_time_str = now.strftime("%Y%m%d_%H%M%S")
# Import custom functions
from my_mods.general import iterate_dict, iterate_list, clear
from my_mods.san import wwpn_colonizer
customer_path = os.path.join(config.customer_path, customer_name)

def main():
    input_directory = os.path.join(customer_path, config.fs_input)
    output_directory = os.path.join(customer_path, config.fs_output)

    try:
        os.listdir(input_directory)
    except:
        print('Error:  Customer folder and/or "fs_input" folder does not exist.  Please create the folder, add the XML files, and re-run the program.')
    else:
        for file_name in os.listdir(input_directory):
            if file_name != '.DS_Store':
                print(file_name)
                full_file = os.path.join(input_directory, file_name)
                tree = ET.ElementTree(file=full_file)
                root = tree.getroot()
                component_list = remove_dup_list(get_components(root))
                component_dict = {}
                for component in component_list:
                    component_dict[component] = xml_to_dict_list(root, component)
                    # if not os.path.exists(f'{output_directory}\{file_name}'):
                    #     os.mkdir(f'{output_directory}\{file_name}')
                    # dict_list_to_csv(component_dict[component], f'{output_directory}\{file_name}\{file_name}-{component}.csv')
                for cluster_dict in component_dict['cluster']:
                    cluster_name = cluster_dict['name']
                    code_level = cluster_dict['code_level']
                    cluster_ip = cluster_dict['console_IP']
                node_count = 0
                for iogrp_dict in component_dict['io_grp']:
                    node_count += int(iogrp_dict['node_count'])
                print(f'NODE COUNT = {node_count}')
                write_to_workbook(output_directory, component_dict, customer_name, cluster_name, code_level, cluster_ip)


def write_to_workbook(output_directory, component_dict, customer_name, cluster_name, code_level, cluster_ip):
    workbook_name = f"{customer_name}_{cluster_name}_QUERYALL_FS_{date_time_str}.xlsx"
    wb = Workbook()
    print(f'\nCreating Workbook: {workbook_name}')
    toc_row = 7
    toc_column = 2
    write_xls_toc(wb, customer_name, cluster_name, toc_row, toc_column, code_level, cluster_ip)
    for key, dict_list in sorted (component_dict.items()):
        toc_row += 1
        write_xls(wb, key, dict_list, toc_row, toc_column)
    if not os.path.exists(output_directory):
        os.makedirs(output_directory)
    wb.save(os.path.join(output_directory,workbook_name))
    print(f"Workbook saved to: {os.path.join(output_directory,workbook_name)}")


def write_xls(wb, key, data, toc_row, toc_column):
    ws = wb.create_sheet(key)
    ws.sheet_view.showGridLines = False
    toc = wb["TOC"]
    toc.cell(row=toc_row, column=toc_column).value = f'=HYPERLINK("#{key}!A1", "{key}")'
    toc.cell(row=toc_row, column=toc_column).font = Font(color=colors.BLUE, underline="single")
    ws.cell(row = 1, column = 1, value = f'=HYPERLINK("#TOC!A1", "Go to Table of Contents")')
    ws.cell(row = 2, column = 1, value = key)
    ws['A1'].font = Font(color=colors.BLUE, italic=True)
    ws['A2'].font = Font(bold=True)
    starting_row = 3
    row = starting_row
    col = 1
    headers = find_max_keys(data)
    
    # This worked in place of using the find_max keys but not ordered correctly
    # # headers = list(set(itertools.chain.from_iterable(data))) 
    
    ws.append(headers)
    for elements in data:
        ws.append([elements.get(h) for h in headers])
    format_xls_table(ws, key, starting_row)
    return wb


'Find the dictionary item with the longest set of keys and return that as list the value'
def find_max_keys(dict_list):
    counter = 0
    old_key = 0
    max_counter = 0 
    for this_dict in dict_list:
        new_key = max_keys(this_dict)
        if new_key and new_key > old_key:
            max_counter = counter
        counter += 1      
    keys = list(dict_list[max_counter].keys())
    return keys


def write_xls_toc(wb, customer_name, cluster_name, toc_row, toc_column, code_level, cluster_ip):
    ws = wb.active
    ws.title = "TOC"
    ws.sheet_view.showGridLines = False
    ws.cell(row = toc_row - 5, column = toc_column, value = customer_name)
    ws.cell(row = toc_row - 4, column = toc_column, value = f'IBM FlashSystem Configuration for cluster:  {cluster_name}')
    ws.cell(row = toc_row - 3, column = toc_column, value = f'Code Level:  {code_level.split()[0]}')
    ws.cell(row = toc_row - 2, column = toc_column, value = f'Cluster IP Address:  {cluster_ip.split(":")[0]}')
    ws.cell(row = toc_row, column = toc_column, value = "Table of Contents")
    ws[f'B{toc_row - 5}'].font = Font(size=18, bold=True)
    ws[f'B{toc_row - 4}'].font = Font(size=14, bold=True)
    ws[f'B{toc_row}'].font = Font(size=14, bold=True)
    return wb


def format_xls_table(ws, key, starting_row):
    last_cell = xlref(ws.max_row,ws.max_column)
    table_name = f'tbl_{key}'
    tab = Table(displayName=table_name, ref= f"A3:{last_cell}")
    # Add a default style with striped rows and banded columns
    style = TableStyleInfo(name="TableStyleMedium1", showFirstColumn=False,
                        showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    tab.tableStyleInfo = style
    table_rows = ws.max_row - starting_row
    if table_rows == 1:
        row_word = "row"
    else:
        row_word = "rows"
    print(f'    Building Table: {table_name} with {table_rows} {row_word}')
    if ws.max_row > starting_row:
        ws.add_table(tab)
    auto_size_columns(ws)
    return ws


def xlref(row, column, zero_indexed=False):
    if zero_indexed:
        row += 1
        column += 1
    return get_column_letter(column) + str(row)


def auto_size_columns(ws):
    dims = {}
    for row in ws.rows:
        for cell in row:
            if cell.value:
                dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))    
    for col, value in dims.items():
        ws.column_dimensions[col].width = value + 5 # Ad


def dict_list_to_csv(dict_list_file, csv_file_name):
    counter = 0
    old_key = 0
    max_counter = 0 
    for this_dict in dict_list_file:
        new_key = max_keys(this_dict)
        if new_key and new_key > old_key:
            max_counter = counter
        counter += 1      
    keys = dict_list_file[max_counter].keys()
    f = open(csv_file_name, 'w', newline='')
    writer = csv.DictWriter(f, fieldnames=keys)
    writer.writeheader()
    for this_dict in dict_list_file:
        writer.writerow(this_dict)
    f.close()


def max_keys(db):
    if db:
        maxcount = max(len(v) for v in db.values())
        return maxcount


def xml_to_dict_list(root, this_type):
    obj_dict_list = []
    for component in root:
        component_type = component.get('type')
        header_dict = {}
        if component_type == this_type:
            attr_dict = {}
            i = 1
            for attr in component:
                attr_name = attr.get('name')
                if attr_name in header_dict.keys():
                    header_dict[attr_name] += 1
                    key = attr_name + str(header_dict[attr_name])
                else:
                    key = attr_name
                    header_dict[attr_name] = 1
                attr_dict[key] = attr.get('value')
            obj_dict_list.append(attr_dict)
    return obj_dict_list


def get_components(root):
    component_list = []
    for component in root:
        component_type = component.get('type')
        component_list.append(component_type)
    return component_list


def remove_dup_list(this_list):
    return list(dict.fromkeys(this_list))


def create_command_list(root):
    command_list = []
    for xclireturn in root:
        command = xclireturn.get('COMMAND_LINE')
        command_list.append(command)
    return command_list


if __name__ == '__main__':
    main()